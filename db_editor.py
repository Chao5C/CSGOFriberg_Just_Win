"""
CS2 Guessr 数据库编辑工具
==========================
功能：查看、添加、编辑、删除、搜索选手
运行：python db_editor.py
"""

import sqlite3
import os
import sys

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "players.db")


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def export_excel():
    """导出数据库到 Excel"""
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT nickname, team, country, age, position,
                   major_wins, major_appearances, status, rating,
                   source, created_at
            FROM players ORDER BY nickname
        """).fetchall()

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "CS2 选手数据库"

        headers = ["昵称", "战队", "国家", "年龄", "位置",
                   "Major冠军", "Major参赛", "状态", "Rating", "来源", "创建时间"]
        hf = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hfont = Font(color="FFFFFF", bold=True, size=11)
        tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

        for col, name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = hfont; cell.fill = hf
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = tb

        for r, rd in enumerate(rows, 2):
            vals = [rd["nickname"], rd["team"], rd["country"], rd["age"],
                    rd["position"], rd["major_wins"], rd["major_appearances"],
                    rd["status"], rd["rating"], rd["source"], rd["created_at"]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = tb
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if r % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

        widths = [16, 18, 14, 8, 12, 12, 12, 10, 10, 10, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:K{len(rows)+1}"

        ws2 = wb.create_sheet("统计")
        team_count, country_count, pos_count = {}, {}, {}
        for rd in rows:
            t, c, p = rd["team"], rd["country"], rd["position"]
            team_count[t] = team_count.get(t, 0) + 1
            country_count[c] = country_count.get(c, 0) + 1
            pos_count[p] = pos_count.get(p, 0) + 1

        ws2.cell(row=1, column=1, value=f"总选手数: {len(rows)}").font = Font(bold=True, size=12)
        ws2.cell(row=3, column=1, value="战队分布").font = Font(bold=True)
        for i, (k, v) in enumerate(sorted(team_count.items(), key=lambda x: -x[1]), 4):
            ws2.cell(row=i, column=1, value=k); ws2.cell(row=i, column=2, value=v)
        ws2.cell(row=3, column=4, value="国家分布").font = Font(bold=True)
        for i, (k, v) in enumerate(sorted(country_count.items(), key=lambda x: -x[1]), 4):
            ws2.cell(row=i, column=4, value=k); ws2.cell(row=i, column=5, value=v)
        ws2.cell(row=3, column=7, value="位置分布").font = Font(bold=True)
        for i, (k, v) in enumerate(sorted(pos_count.items(), key=lambda x: -x[1]), 4):
            ws2.cell(row=i, column=7, value=k); ws2.cell(row=i, column=8, value=v)
        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['D'].width = 16
        ws2.column_dimensions['G'].width = 16

        path = os.path.join(DATA_DIR, "players.xlsx")
        wb.save(path)
        return len(rows)
    except Exception as e:
        print(f"导出失败: {e}")
        return 0
    finally:
        conn.close()


# ============ 增删改查 ============

def list_all():
    conn = get_db()
    rows = conn.execute("SELECT * FROM players ORDER BY nickname").fetchall()
    conn.close()
    if not rows:
        print("\n  数据库为空")
        return rows
    print(f"\n  {'ID':<5} {'昵称':<20} {'战队':<18} {'国家':<14} {'年龄':<5} {'位置':<10} {'Major冠':<8} {'Major赛':<8} {'状态':<8} {'Rating':<7} {'来源':<8}")
    print("  " + "-" * 120)
    for r in rows:
        print(f"  {r['id']:<5} {r['nickname']:<20} {r['team']:<18} {r['country']:<14} {r['age']:<5} "
              f"{r['position']:<10} {r['major_wins']:<8} {r['major_appearances']:<8} "
              f"{r['status']:<8} {r['rating']:<7.2f} {r['source']:<8}")
    print(f"\n  共 {len(rows)} 位选手")
    return rows


def search_player(keyword):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM players WHERE nickname LIKE ? OR team LIKE ? OR country LIKE ? ORDER BY nickname",
        (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%")
    ).fetchall()
    conn.close()
    if not rows:
        print(f"\n  未找到匹配 '{keyword}' 的选手")
        return rows
    print(f"\n  搜索 '{keyword}' 结果 ({len(rows)} 条):")
    for r in rows:
        print(f"  [{r['id']}] {r['nickname']} | {r['team']} | {r['country']} | "
              f"年龄={r['age']} | {r['position']} | Majors: {r['major_wins']}/{r['major_appearances']} | "
              f"{r['status']} | Rating={r['rating']}")
    return rows


def add_player():
    print("\n  === 添加选手 ===")
    nickname = input("  昵称: ").strip()
    if not nickname:
        print("  昵称不能为空！")
        return
    team = input("  战队: ").strip() or "Unknown"
    country = input("  国家: ").strip() or "Unknown"
    age_str = input("  年龄: ").strip()
    try:
        age = int(age_str) if age_str else 0
    except:
        age = 0
    position = input("  位置 (Rifler/AWPer/IGL/Support): ").strip() or "Rifler"
    mw_str = input("  Major冠军数: ").strip()
    try:
        mw = int(mw_str) if mw_str else 0
    except:
        mw = 0
    ma_str = input("  Major参赛数: ").strip()
    try:
        ma = int(ma_str) if ma_str else 0
    except:
        ma = 0
    status = input("  状态 (Active/Inactive/Retired): ").strip() or "Active"
    r_str = input("  Rating: ").strip()
    try:
        rating = float(r_str) if r_str else 0.0
    except:
        rating = 0.0

    conn = get_db()
    try:
        conn.execute("""
            INSERT OR REPLACE INTO players
            (nickname,team,country,age,position,major_wins,major_appearances,status,rating,source)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (nickname, team, country, age, position, mw, ma, status, rating, "manual"))
        conn.commit()
        print(f"\n  ✅ 已添加/更新: {nickname}")
        export_excel()
    except Exception as e:
        print(f"\n  ❌ 添加失败: {e}")
    finally:
        conn.close()


def edit_player():
    print("\n  === 编辑选手 ===")
    nickname = input("  选手昵称: ").strip()
    if not nickname:
        print("  昵称不能为空！")
        return

    conn = get_db()
    row = conn.execute("SELECT * FROM players WHERE nickname LIKE ?",
                       (f"%{nickname}%",)).fetchone()
    if not row:
        print(f"  未找到匹配 '{nickname}' 的选手")
        conn.close()
        return

    pid = row["id"]
    fields = {
        "昵称": ("nickname", row["nickname"]),
        "战队": ("team", row["team"]),
        "国家": ("country", row["country"]),
        "年龄": ("age", str(row["age"])),
        "位置": ("position", row["position"]),
        "Major冠军数": ("major_wins", str(row["major_wins"])),
        "Major参赛数": ("major_appearances", str(row["major_appearances"])),
        "状态": ("status", row["status"]),
        "Rating": ("rating", str(row["rating"])),
    }

    print(f"\n  当前数据 [{pid}] {row['nickname']} | {row['team']} | {row['country']}:")
    print(f"  逐项回车保留原值，输入 'q' 取消\n")
    for label, (col, val) in fields.items():
        new_val = input(f"  {label} [{val}]: ").strip()
        if new_val.lower() == 'q':
            print("  已取消")
            conn.close()
            return
        if new_val:
            if col in ("age", "major_wins", "major_appearances"):
                try:
                    new_val = int(new_val)
                except:
                    continue
            elif col == "rating":
                try:
                    new_val = float(new_val)
                except:
                    continue
            conn.execute(f"UPDATE players SET {col}=? WHERE id=?", (new_val, pid))

    conn.commit()
    conn.close()
    print(f"\n  ✅ 已更新选手 [{pid}] {row['nickname']}")
    export_excel()


def delete_player():
    print("\n  === 删除选手 ===")
    nickname = input("  选手昵称: ").strip()
    if not nickname:
        print("  昵称不能为空！")
        return

    conn = get_db()
    row = conn.execute("SELECT * FROM players WHERE nickname LIKE ?",
                       (f"%{nickname}%",)).fetchone()
    if not row:
        print(f"  未找到匹配 '{nickname}' 的选手")
        conn.close()
        return

    confirm = input(f"\n  确认删除 [{row['id']}] {row['nickname']} ({row['team']})? [y/N]: ").strip()
    if confirm.lower() == 'y':
        conn.execute("DELETE FROM players WHERE id=?", (row["id"],))
        conn.commit()
        print(f"  ✅ 已删除: {row['nickname']}")
        export_excel()
    else:
        print("  已取消")
    conn.close()


def show_stats():
    conn = get_db()
    total = conn.execute("SELECT COUNT(*) as c FROM players").fetchone()["c"]
    teams = conn.execute("""
        SELECT team, COUNT(*) as c FROM players
        GROUP BY team ORDER BY c DESC LIMIT 10
    """).fetchall()
    countries = conn.execute("""
        SELECT country, COUNT(*) as c FROM players
        GROUP BY country ORDER BY c DESC LIMIT 10
    """).fetchall()
    positions = conn.execute("""
        SELECT position, COUNT(*) as c FROM players
        GROUP BY position ORDER BY c DESC
    """).fetchall()
    avg_rating = conn.execute("""
        SELECT AVG(rating) as avg, MAX(rating) as mx, MIN(rating) as mn
        FROM players WHERE rating > 0
    """).fetchone()
    statuses = conn.execute("""
        SELECT status, COUNT(*) as c FROM players GROUP BY status
    """).fetchall()
    conn.close()

    print(f"\n  {'='*50}")
    print(f"  数据库统计")
    print(f"  {'='*50}")
    print(f"  总选手数: {total}")
    if avg_rating and avg_rating["avg"]:
        print(f"  Rating: 平均 {avg_rating['avg']:.2f} | 最高 {avg_rating['mx']:.2f} | 最低 {avg_rating['mn']:.2f}")

    print(f"\n  状态分布:")
    for s in statuses:
        print(f"    {s['status']}: {s['c']}")

    print(f"\n  位置分布:")
    for p in positions:
        print(f"    {p['position']}: {p['c']}")

    print(f"\n  战队 TOP 10:")
    for t in teams:
        print(f"    {t['team']:<25} {t['c']}")

    print(f"\n  国家 TOP 10:")
    for c in countries:
        print(f"    {c['country']:<20} {c['c']}")


def batch_add():
    """批量添加 - 从剪贴板或输入多行"""
    print(f"\n  === 批量添加选手 ===")
    print(f"  格式: 昵称 | 战队 | 国家 | 年龄 | 位置 | Major冠军 | Major参赛 | 状态 | Rating")
    print(f"  示例: s1mple | NAVI | Ukraine | 26 | AWPer | 1 | 10 | Active | 1.25")
    print(f"  每行一个选手，空行结束\n")

    lines = []
    while True:
        line = input("  > ").strip()
        if not line:
            break
        lines.append(line)

    if not lines:
        print("  取消")
        return

    added = 0
    for line in lines:
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 3:
            print(f"  跳过(格式不对): {line}")
            continue

        nickname = parts[0]
        team = parts[1] if len(parts) > 1 else "Unknown"
        country = parts[2] if len(parts) > 2 else "Unknown"
        age = int(parts[3]) if len(parts) > 3 and parts[3] else 0
        position = parts[4] if len(parts) > 4 and parts[4] else "Rifler"
        mw = int(parts[5]) if len(parts) > 5 and parts[5] else 0
        ma = int(parts[6]) if len(parts) > 6 and parts[6] else 0
        status = parts[7] if len(parts) > 7 and parts[7] else "Active"
        rating = float(parts[8]) if len(parts) > 8 and parts[8] else 0.0

        conn = get_db()
        try:
            conn.execute("""
                INSERT OR REPLACE INTO players
                (nickname,team,country,age,position,major_wins,major_appearances,status,rating,source)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (nickname, team, country, age, position, mw, ma, status, rating, "manual"))
            conn.commit()
            print(f"  ✅ {nickname}")
            added += 1
        except Exception as e:
            print(f"  ❌ {nickname}: {e}")
        finally:
            conn.close()

    if added:
        export_excel()
        print(f"\n  共添加 {added} 位选手")


# ============ 菜单 ============

def menu():
    while True:
        print("\n")
        print("  " + "=" * 50)
        print("    CS2 Guessr 数据库编辑工具")
        print("  " + "=" * 50)
        print("    1. 查看所有选手")
        print("    2. 搜索选手")
        print("    3. 添加选手")
        print("    4. 编辑选手")
        print("    5. 删除选手")
        print("    6. 批量添加")
        print("    7. 统计信息")
        print("    8. 导出 Excel")
        print("    0. 退出")
        print("  " + "-" * 50)

        try:
            choice = input("  请选择: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n  再见！")
            break

        if choice == "1":
            list_all()
        elif choice == "2":
            kw = input("\n  搜索关键词: ").strip()
            if kw:
                search_player(kw)
        elif choice == "3":
            add_player()
        elif choice == "4":
            edit_player()
        elif choice == "5":
            delete_player()
        elif choice == "6":
            batch_add()
        elif choice == "7":
            show_stats()
        elif choice == "8":
            count = export_excel()
            print(f"\n  ✅ 已导出 {count} 位选手到 data/players.xlsx")
        elif choice == "0":
            print("\n  再见！")
            break
        else:
            print(f"\n  无效选项: {choice}")


if __name__ == "__main__":
    os.makedirs(DATA_DIR, exist_ok=True)
    menu()
