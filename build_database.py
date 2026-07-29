"""
数据库构建脚本
- 创建 SQLite 数据库
- 导入爬虫数据 / 手动添加选手
- 导出为 Excel 文件
"""
import os
import sqlite3

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
DB_PATH = os.path.join(DATA_DIR, "players.db")
EXCEL_PATH = os.path.join(DATA_DIR, "players.xlsx")


def ensure_dir():
    os.makedirs(DATA_DIR, exist_ok=True)


def get_conn():
    """获取数据库连接"""
    ensure_dir()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


# ============ 建表 ============

def create_tables():
    """创建数据库表"""
    conn = get_conn()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS players (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nickname TEXT NOT NULL UNIQUE,
            team TEXT NOT NULL,
            country TEXT NOT NULL,
            age INTEGER NOT NULL DEFAULT 0,
            position TEXT NOT NULL DEFAULT 'Rifler',
            major_wins INTEGER NOT NULL DEFAULT 0,
            major_appearances INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'Active',
            rating REAL DEFAULT 0,
            source TEXT DEFAULT 'manual',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()
    print("[数据库] 表已就绪")


# ============ 选手 CRUD ============

def add_player(nickname: str, team: str, country: str, age: int,
               position: str = "Rifler", major_wins: int = 0,
               major_appearances: int = 0, status: str = "Active",
               rating: float = 0, source: str = "manual") -> bool:
    """添加一个选手，存在则更新"""
    conn = get_conn()
    try:
        conn.execute("""
            INSERT INTO players (nickname, team, country, age, position,
                                 major_wins, major_appearances, status, rating, source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(nickname) DO UPDATE SET
                team=excluded.team, country=excluded.country, age=excluded.age,
                position=excluded.position, major_wins=excluded.major_wins,
                major_appearances=excluded.major_appearances, status=excluded.status,
                rating=excluded.rating, source=excluded.source
        """, (nickname, team, country, age, position,
              major_wins, major_appearances, status, rating, source))
        conn.commit()
        return True
    except Exception as e:
        print(f"[数据库] 添加失败 {nickname}: {e}")
        return False
    finally:
        conn.close()


def add_players_batch(players: list) -> int:
    """批量添加选手"""
    added = 0
    for p in players:
        if add_player(
            nickname=p.get("nickname", ""),
            team=p.get("team", "Unknown"),
            country=p.get("country", "Unknown"),
            age=p.get("age", 0),
            position=p.get("position", "Rifler"),
            major_wins=p.get("major_wins", 0),
            major_appearances=p.get("major_appearances", 0),
            status=p.get("status", "Active"),
            rating=p.get("rating", 0),
            source=p.get("source", "manual"),
        ):
            added += 1
    return added


def get_all_players() -> list:
    """获取全部选手"""
    conn = get_conn()
    try:
        rows = conn.execute("SELECT * FROM players ORDER BY nickname").fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def get_player_count() -> int:
    conn = get_conn()
    try:
        return conn.execute("SELECT COUNT(*) as c FROM players").fetchone()["c"]
    finally:
        conn.close()


def delete_player(nickname: str) -> bool:
    conn = get_conn()
    try:
        conn.execute("DELETE FROM players WHERE nickname = ?", (nickname,))
        conn.commit()
        return True
    finally:
        conn.close()


def clear_all():
    """清空数据库"""
    conn = get_conn()
    conn.execute("DELETE FROM players")
    conn.commit()
    conn.close()
    print("[数据库] 已清空")


# ============ Excel 导出 ============

def export_to_excel(output_path: str = None) -> str:
    """导出全部选手到 Excel"""
    if output_path is None:
        output_path = EXCEL_PATH

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[导出] openpyxl 未安装，请执行: pip install openpyxl")
        return ""

    ensure_dir()
    players = get_all_players()

    wb = Workbook()
    ws = wb.active
    ws.title = "CS2 Players"

    # 表头
    headers = ["ID", "昵称", "队伍", "国家/地区", "年龄", "位置",
               "Major冠军", "Major次数", "状态", "Rating", "来源", "创建时间"]
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    data_font = Font(name="Microsoft YaHei", size=10)
    data_align = Alignment(horizontal="center", vertical="center")
    # 状态颜色
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_idx, p in enumerate(players, 2):
        values = [p["id"], p["nickname"], p["team"], p["country"],
                  p["age"], p["position"], p["major_wins"],
                  p["major_appearances"], p["status"],
                  p["rating"], p["source"],
                  p["created_at"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

        # 状态列着色
        status_cell = ws.cell(row=row_idx, column=9)
        if p["status"] == "Active":
            status_cell.fill = green_fill
        elif p["status"] == "Inactive":
            status_cell.fill = yellow_fill
        elif p["status"] == "Retired":
            status_cell.fill = red_fill

    # 列宽
    col_widths = [5, 20, 18, 16, 6, 12, 10, 10, 10, 8, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"[导出] Excel 已保存至 {output_path} ({len(players)} 条记录)")
    return output_path


def export_to_sql(output_path: str = None) -> str:
    """导出为 SQL 文件"""
    if output_path is None:
        output_path = os.path.join(DATA_DIR, "players_export.sql")

    ensure_dir()
    players = get_all_players()

    lines = [
        "-- CS2 Players Database Export",
        f"-- 导出时间: {__import__('datetime').datetime.now().isoformat()}",
        f"-- 记录数: {len(players)}",
        "",
        "BEGIN TRANSACTION;",
        "",
    ]

    for p in players:
        nick = p['nickname'].replace("'", "''")
        team = p['team'].replace("'", "''")
        country = p['country'].replace("'", "''")
        pos = p['position'].replace("'", "''")
        stat = p['status'].replace("'", "''")
        src = p['source'].replace("'", "''")
        lines.append(
            "INSERT OR REPLACE INTO players "
            "(nickname, team, country, age, position, major_wins, major_appearances, status, rating, source) "
            "VALUES ("
            f"'{nick}', '{team}', '{country}', {p['age']}, "
            f"'{pos}', {p['major_wins']}, {p['major_appearances']}, "
            f"'{stat}', {p['rating']}, '{src}'"
            ");"
        )

    lines.append("")
    lines.append("COMMIT;")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"[导出] SQL 已保存至 {output_path}")
    return output_path


# ============ 种子数据 ============

SEED_DATA = [
    # Spirit
    ("donk", "Spirit", "Russia", 18, "Rifler", 1, 2, "Active", 1.31),
    ("sh1ro", "Spirit", "Russia", 23, "AWPer", 0, 5, "Active", 1.18),
    ("magixx", "Spirit", "Russia", 22, "Rifler", 1, 4, "Active", 1.04),
    ("zont1x", "Spirit", "Ukraine", 20, "Rifler", 1, 2, "Active", 1.10),
    ("chopper", "Spirit", "Russia", 28, "IGL", 1, 8, "Active", 0.93),
    # Vitality
    ("ZywOo", "Vitality", "France", 24, "AWPer", 1, 6, "Active", 1.29),
    ("flameZ", "Vitality", "Israel", 22, "Rifler", 0, 4, "Active", 1.11),
    ("apEX", "Vitality", "France", 31, "IGL", 2, 15, "Active", 0.97),
    ("mezii", "Vitality", "UK", 26, "Rifler", 0, 5, "Active", 1.03),
    ("ropz", "Vitality", "Estonia", 25, "Rifler", 1, 9, "Active", 1.13),
    # G2
    ("m0NESY", "G2", "Russia", 19, "AWPer", 0, 3, "Active", 1.24),
    ("NiKo", "G2", "Bosnia", 28, "Rifler", 0, 10, "Active", 1.19),
    ("huNter-", "G2", "Bosnia", 29, "Rifler", 0, 8, "Active", 1.07),
    ("malbsMd", "G2", "Guatemala", 21, "Rifler", 0, 2, "Active", 1.09),
    ("Snax", "G2", "Poland", 31, "IGL", 1, 14, "Active", 0.93),
    # NAVI
    ("b1t", "NAVI", "Ukraine", 22, "Rifler", 1, 4, "Active", 1.12),
    ("jL", "NAVI", "Lithuania", 25, "Rifler", 1, 4, "Active", 1.13),
    ("iM", "NAVI", "Romania", 25, "Rifler", 1, 5, "Active", 1.06),
    ("Aleksib", "NAVI", "Finland", 28, "IGL", 1, 8, "Active", 0.97),
    ("w0nderful", "NAVI", "Ukraine", 20, "AWPer", 1, 2, "Active", 1.07),
    # FaZe
    ("broky", "FaZe", "Latvia", 23, "AWPer", 1, 5, "Active", 1.11),
    ("frozen", "FaZe", "Slovakia", 22, "Rifler", 0, 6, "Active", 1.12),
    ("rain", "FaZe", "Norway", 30, "Rifler", 1, 12, "Active", 1.03),
    ("karrigan", "FaZe", "Denmark", 34, "IGL", 1, 18, "Active", 0.90),
    ("elige", "FaZe", "USA", 27, "Rifler", 0, 8, "Active", 1.08),
    # MOUZ
    ("Jimpphat", "MOUZ", "Finland", 18, "Rifler", 0, 1, "Active", 1.12),
    ("torzsi", "MOUZ", "Hungary", 22, "AWPer", 0, 3, "Active", 1.10),
    ("xertioN", "MOUZ", "Israel", 20, "Rifler", 0, 2, "Active", 1.07),
    ("Brollan", "MOUZ", "Sweden", 22, "IGL", 0, 5, "Active", 1.03),
    ("siuhy", "MOUZ", "Poland", 22, "IGL", 0, 4, "Active", 0.96),
    # Eternal Fire
    ("XANTARES", "Eternal Fire", "Turkey", 29, "Rifler", 0, 9, "Active", 1.21),
    ("woxic", "Eternal Fire", "Turkey", 26, "AWPer", 0, 6, "Active", 1.10),
    ("MAJ3R", "Eternal Fire", "Turkey", 34, "IGL", 0, 14, "Active", 0.98),
    # Falcons
    ("s1mple", "Falcons", "Ukraine", 27, "AWPer", 1, 10, "Active", 1.24),
    ("NiKo", "Falcons", "Bosnia", 28, "Rifler", 0, 10, "Active", 1.19),
    ("Magisk", "Falcons", "Denmark", 27, "Rifler", 4, 10, "Active", 1.05),
    # Virtus.pro
    ("FL1T", "Virtus.pro", "Russia", 24, "Rifler", 0, 5, "Active", 1.08),
    ("Jame", "Virtus.pro", "Russia", 26, "IGL", 1, 8, "Active", 1.06),
    ("fame", "Virtus.pro", "Russia", 22, "Rifler", 1, 2, "Active", 1.06),
    ("electroNic", "Virtus.pro", "Russia", 26, "Rifler", 0, 7, "Active", 1.05),
    # Liquid
    ("Twistzz", "Liquid", "Canada", 25, "Rifler", 2, 8, "Active", 1.14),
    ("NAF", "Liquid", "Canada", 27, "Rifler", 1, 9, "Active", 1.08),
    ("jks", "Liquid", "Australia", 29, "Rifler", 1, 9, "Active", 1.02),
    ("ultimate", "Liquid", "Poland", 21, "AWPer", 0, 2, "Active", 1.04),
    # FURIA
    ("KSCERATO", "FURIA", "Brazil", 25, "Rifler", 0, 6, "Active", 1.16),
    ("yuurih", "FURIA", "Brazil", 25, "Rifler", 0, 7, "Active", 1.06),
    ("FalleN", "FURIA", "Brazil", 33, "IGL", 2, 18, "Active", 0.98),
    ("chelo", "FURIA", "Brazil", 26, "Rifler", 0, 5, "Active", 1.04),
    # Astralis
    ("device", "Astralis", "Denmark", 29, "AWPer", 4, 12, "Active", 1.16),
    ("stavn", "Astralis", "Denmark", 23, "Rifler", 0, 4, "Active", 1.10),
    ("jabbi", "Astralis", "Denmark", 21, "Rifler", 0, 3, "Active", 1.07),
    ("cadiaN", "Astralis", "Denmark", 29, "IGL", 0, 10, "Active", 1.02),
    # HEROIC
    ("kyxsan", "HEROIC", "North Macedonia", 25, "IGL", 0, 5, "Active", 0.96),
    ("NertZ", "HEROIC", "Israel", 25, "Rifler", 0, 4, "Active", 1.11),
    ("SunPayus", "HEROIC", "Spain", 26, "AWPer", 0, 5, "Active", 1.08),
    # BIG
    ("tabseN", "BIG", "Germany", 29, "IGL", 0, 9, "Active", 1.10),
    ("rigoN", "BIG", "Switzerland", 25, "Rifler", 0, 5, "Active", 1.07),
    # Complexity
    ("hallzerk", "Complexity", "Norway", 24, "AWPer", 0, 4, "Active", 1.05),
    ("Grim", "Complexity", "USA", 24, "Rifler", 0, 5, "Active", 1.03),
    # ENCE
    ("gla1ve", "ENCE", "Denmark", 29, "IGL", 4, 14, "Active", 0.91),
    ("sdy", "ENCE", "Ukraine", 27, "Rifler", 0, 7, "Active", 1.05),
    ("Dycha", "ENCE", "Poland", 27, "Rifler", 0, 6, "Active", 1.03),
    # Cloud9
    ("Ax1Le", "Cloud9", "Russia", 22, "Rifler", 0, 3, "Active", 1.09),
    ("Boombl4", "Cloud9", "Russia", 26, "IGL", 1, 7, "Active", 0.98),
    ("Perfecto", "Cloud9", "Russia", 25, "Rifler", 1, 6, "Active", 1.01),
    # M80 / Wildcard
    ("swisher", "M80", "USA", 24, "Rifler", 0, 3, "Active", 1.08),
    # SAW
    ("story", "SAW", "Portugal", 25, "IGL", 0, 6, "Active", 0.98),
    ("ewjerkz", "SAW", "Portugal", 23, "Rifler", 0, 3, "Active", 1.07),
    # Imperial fe
    ("ANa", "Imperial fe", "Russia", 26, "AWPer", 0, 2, "Active", 1.12),
    # 3DMAX
    ("Maka", "3DMAX", "France", 25, "IGL", 0, 6, "Active", 1.01),
    # TYLOO + 亚洲
    ("xccurate", "TYLOO", "Indonesia", 27, "AWPer", 0, 6, "Active", 1.07),
    ("JamYoung", "TYLOO", "China", 24, "Rifler", 0, 4, "Active", 1.05),
    ("Mercury", "TYLOO", "China", 21, "Rifler", 0, 2, "Active", 1.03),
    ("somebody", "TYLOO", "China", 29, "Rifler", 0, 8, "Active", 0.98),
    ("Attacker", "TYLOO", "China", 28, "IGL", 0, 7, "Active", 0.95),
    ("aumaN", "Rare Atom", "China", 23, "Rifler", 0, 2, "Active", 1.06),
    ("kaze", "Rare Atom", "Malaysia", 30, "AWPer", 0, 10, "Active", 1.09),
    # 已退役知名选手
    ("olofmeister", "Retired", "Sweden", 32, "Rifler", 2, 12, "Retired", 1.02),
    ("dupreeh", "Retired", "Denmark", 31, "Rifler", 5, 18, "Retired", 0.98),
    ("Xyp9x", "Retired", "Denmark", 29, "Support", 4, 15, "Retired", 0.95),
    ("GuardiaN", "Retired", "Slovakia", 33, "AWPer", 0, 10, "Retired", 0.92),
    ("kennyS", "Retired", "France", 29, "AWPer", 1, 8, "Retired", 0.96),
    ("GeT_RiGhT", "Retired", "Sweden", 34, "Rifler", 0, 12, "Retired", 0.88),
    ("f0rest", "Retired", "Sweden", 36, "Rifler", 0, 15, "Retired", 0.90),
    ("shox", "Retired", "France", 32, "Rifler", 0, 12, "Retired", 0.94),
    ("NBK-", "Retired", "France", 30, "Rifler", 2, 11, "Retired", 0.93),
    ("JW", "Retired", "Sweden", 30, "AWPer", 3, 10, "Retired", 0.89),
    ("flusha", "Retired", "Sweden", 31, "Rifler", 3, 11, "Retired", 0.91),
    ("coldzera", "Retired", "Brazil", 30, "Rifler", 2, 8, "Retired", 0.95),
    ("fer", "Retired", "Brazil", 33, "Rifler", 2, 9, "Retired", 0.88),
    ("TACO", "Retired", "Brazil", 29, "Support", 2, 8, "Retired", 0.85),
]


def seed_data():
    """导入种子数据"""
    added = 0
    for p in SEED_DATA:
        nickname, team, country, age, position, major_wins, major_appearances, status, rating = p
        if add_player(nickname, team, country, age, position, major_wins,
                      major_appearances, status, rating, "seed"):
            added += 1
    print(f"[种子数据] 导入 {added} 位选手")


# ============ 主入口 ============

def build(seed: bool = True):
    """
    构建/更新数据库
    1. 创建表
    2. 导入种子数据（可选）
    3. 导出 Excel + SQL
    """
    ensure_dir()
    create_tables()

    if seed:
        seed_data()

    count = get_player_count()
    print(f"\n[数据库] 当前共 {count} 位选手")

    # 导出
    export_to_excel()
    export_to_sql()

    print(f"\n[完成] 数据库: {DB_PATH}")
    print(f"[完成] Excel:  {EXCEL_PATH}")
    return count


if __name__ == "__main__":
    import sys
    do_seed = "--no-seed" not in sys.argv
    build(seed=do_seed)
