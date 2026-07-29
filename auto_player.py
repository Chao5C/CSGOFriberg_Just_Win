"""
CS2 Guessr 自动游戏脚本 (v2)
=============================
核心改进：
- 通过网络拦截 API 响应获取答案（而非 DOM 解析）
- 捕获游戏初始化 / 猜测验证 / 答案揭示的所有 API 数据
- 自动提取选手完整信息并存入数据库
- 每轮截图辅助调试

运行: python auto_player.py [轮数]
"""

import asyncio
import json
import os
import sys
import sqlite3
import re
import time
from datetime import datetime
from typing import Optional

from playwright.async_api import async_playwright, Page

# ============ 路径 ============
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "players.db")
EXCEL_PATH = os.path.join(DATA_DIR, "players.xlsx")
SCREENSHOT_DIR = os.path.join(DATA_DIR, "screenshots")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

# ============ 地区映射 ============
COUNTRY_REGION_MAP = {
    "Denmark": "Europe","Sweden": "Europe","Norway": "Europe","Finland": "Europe",
    "France": "Europe","Germany": "Europe","UK": "Europe","United Kingdom": "Europe",
    "Netherlands": "Europe","Poland": "Europe","Spain": "Europe","Portugal": "Europe",
    "Belgium": "Europe","Switzerland": "Europe","Austria": "Europe","Estonia": "Europe",
    "Latvia": "Europe","Lithuania": "Europe","Czech Republic": "Europe","Slovakia": "Europe",
    "Hungary": "Europe","Romania": "Europe","Bulgaria": "Europe","Serbia": "Europe",
    "Croatia": "Europe","Slovenia": "Europe","Bosnia and Herzegovina": "Europe",
    "Bosnia": "Europe","Montenegro": "Europe","North Macedonia": "Europe","Kosovo": "Europe",
    "Turkey": "Europe","Israel": "Europe","Italy": "Europe","Greece": "Europe",
    "Iceland": "Europe","Ireland": "Europe","Malta": "Europe","Luxembourg": "Europe",
    "Scotland": "Europe","Wales": "Europe","England": "Europe",
    "Russia": "CIS","Ukraine": "CIS","Belarus": "CIS","Kazakhstan": "CIS",
    "Uzbekistan": "CIS","Armenia": "CIS","Azerbaijan": "CIS","Georgia": "CIS",
    "Moldova": "CIS","Kyrgyzstan": "CIS",
    "USA": "Americas","United States": "Americas","Canada": "Americas","Brazil": "Americas",
    "Argentina": "Americas","Chile": "Americas","Mexico": "Americas","Colombia": "Americas",
    "Peru": "Americas","Uruguay": "Americas","Venezuela": "Americas","Ecuador": "Americas",
    "Guatemala": "Americas","Dominican Republic": "Americas","Puerto Rico": "Americas",
    "China": "Asia","Mongolia": "Asia","South Korea": "Asia","Japan": "Asia",
    "Taiwan": "Asia","Vietnam": "Asia","Thailand": "Asia","Indonesia": "Asia",
    "Malaysia": "Asia","Philippines": "Asia","India": "Asia","Singapore": "Asia",
    "Hong Kong": "Asia","Macau": "Asia",
    "Australia": "Oceania","New Zealand": "Oceania",
    "South Africa": "Africa","Egypt": "Africa","Morocco": "Africa",
    "Tunisia": "Africa","Algeria": "Africa",
}

POSITION_GROUPS = {
    "AWPer": ["awper","sniper","狙击手"],
    "IGL": ["igl","captain","in-game leader","指挥"],
    "Rifler": ["rifler","entry","entry fragger","lurker","anchor","fragger","步枪手","突破"],
    "Support": ["support","supportive","辅助"],
    "Coach": ["coach","教练"],
}


# ============ 数据库 ============
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS players (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nickname TEXT NOT NULL UNIQUE,
            team TEXT NOT NULL,
            country TEXT NOT NULL,
            age INTEGER NOT NULL DEFAULT 0,
            position TEXT NOT NULL DEFAULT 'Rifler',
            major_wins INTEGER NOT NULL DEFAULT 0,
            major_appearances INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'Active',
            rating REAL DEFAULT 0,
            source TEXT DEFAULT 'auto',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

def add_player(nickname, team, country, age, position="Rifler",
               major_wins=0, major_appearances=0, status="Active",
               rating=0.0, source="auto"):
    conn = get_db()
    try:
        conn.execute("""
            INSERT OR REPLACE INTO players
            (nickname,team,country,age,position,major_wins,major_appearances,status,rating,source)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (nickname, team, country, age, position, major_wins,
              major_appearances, status, rating, source))
        conn.commit()
        # 实时同步到 Excel
        export_to_excel()
        return True
    except Exception as e:
        print(f"  [DB] 添加失败 {nickname}: {e}")
        return False
    finally:
        conn.close()

def export_to_excel():
    """实时将数据库导出到 Excel 文件"""
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT nickname, team, country, age, position,
                   major_wins, major_appearances, status, rating,
                   source, created_at
            FROM players
            ORDER BY nickname
        """).fetchall()

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "CS2 选手数据库"

        # 表头
        headers = ["昵称", "战队", "国家", "年龄", "位置",
                   "Major冠军", "Major参赛", "状态", "Rating", "来源", "创建时间"]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # 数据行
        for r, row_data in enumerate(rows, 2):
            values = [
                row_data["nickname"], row_data["team"], row_data["country"],
                row_data["age"], row_data["position"],
                row_data["major_wins"], row_data["major_appearances"],
                row_data["status"], row_data["rating"],
                row_data["source"], row_data["created_at"]
            ]
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # 交替行颜色
                if r % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

        # 列宽
        widths = [16, 18, 14, 8, 12, 12, 12, 10, 10, 10, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # 冻结首行
        ws.freeze_panes = "A2"
        # 自动筛选
        ws.auto_filter.ref = f"A1:K{len(rows)+1}"

        # 添加统计 sheet
        ws2 = wb.create_sheet("统计")
        team_count = {}
        country_count = {}
        pos_count = {}
        for row_data in rows:
            t = row_data["team"]; team_count[t] = team_count.get(t, 0) + 1
            c = row_data["country"]; country_count[c] = country_count.get(c, 0) + 1
            p = row_data["position"]; pos_count[p] = pos_count.get(p, 0) + 1

        ws2.cell(row=1, column=1, value=f"总选手数: {len(rows)}").font = Font(bold=True, size=12)
        ws2.cell(row=3, column=1, value="战队分布").font = Font(bold=True)
        for i, (t, c) in enumerate(sorted(team_count.items(), key=lambda x: -x[1]), 4):
            ws2.cell(row=i, column=1, value=t)
            ws2.cell(row=i, column=2, value=c)

        ws2.cell(row=3, column=4, value="国家分布").font = Font(bold=True)
        for i, (t, c) in enumerate(sorted(country_count.items(), key=lambda x: -x[1]), 4):
            ws2.cell(row=i, column=4, value=t)
            ws2.cell(row=i, column=5, value=c)

        ws2.cell(row=3, column=7, value="位置分布").font = Font(bold=True)
        for i, (t, c) in enumerate(sorted(pos_count.items(), key=lambda x: -x[1]), 4):
            ws2.cell(row=i, column=7, value=t)
            ws2.cell(row=i, column=8, value=c)

        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['D'].width = 16
        ws2.column_dimensions['G'].width = 16

        wb.save(EXCEL_PATH)
        return len(rows)
    except Exception as e:
        print(f"  [Excel] 导出失败: {e}")
        return 0
    finally:
        conn.close()


def load_all_players():
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM players ORDER BY nickname").fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()

def get_player_count():
    conn = get_db()
    try:
        return conn.execute("SELECT COUNT(*) as c FROM players").fetchone()["c"]
    finally:
        conn.close()

def get_region(country):
    return COUNTRY_REGION_MAP.get(country, "Other")

def get_position_group(pos):
    pl = (pos or "").lower()
    for group, variants in POSITION_GROUPS.items():
        for v in variants:
            if v in pl:
                return group
    return "Rifler"


# ============ 候选筛选 ============
def filter_candidates(candidates, guess_player, feedbacks):
    filtered = []
    for cand in candidates:
        if cand["nickname"].lower() == (guess_player.get("nickname","") or "").lower():
            continue
        match = True

        # 队伍
        fb = feedbacks.get("team", {})
        gt = (guess_player.get("team","") or "")
        ct = (cand.get("team","") or "")
        if fb.get("color") == "green":
            if ct.lower() != gt.lower(): match = False
        elif fb.get("color") in ("gray","wrong"):
            if ct.lower() == gt.lower(): match = False
        if not match: continue

        # 国家
        fb = feedbacks.get("country", {})
        gc = (guess_player.get("country","") or "")
        cc = (cand.get("country","") or "")
        if fb.get("color") == "green":
            if cc.lower() != gc.lower(): match = False
        elif fb.get("color") in ("yellow","close"):
            if get_region(cc) != get_region(gc): match = False
        elif fb.get("color") in ("gray","wrong"):
            if cc.lower() == gc.lower(): match = False
            elif get_region(cc) == get_region(gc): match = False
        if not match: continue

        # 年龄
        fb = feedbacks.get("age", {})
        ga = guess_player.get("age", 0) or 0
        ca = cand.get("age", 0) or 0
        if fb.get("color") == "green":
            if ca != ga: match = False
        elif fb.get("color") in ("yellow","close"):
            if abs(ca - ga) > 2: match = False
        elif fb.get("color") in ("gray","wrong"):
            if abs(ca - ga) <= 2: match = False
        if match and fb.get("direction"):
            if fb["direction"] == "higher" and ca >= ga: match = False
            if fb["direction"] == "lower" and ca <= ga: match = False
        if not match: continue

        # 位置
        fb = feedbacks.get("position", {})
        gp = (guess_player.get("position","") or "")
        cp = (cand.get("position","") or "")
        if fb.get("color") == "green":
            if cp.lower() != gp.lower(): match = False
        elif fb.get("color") in ("yellow","close"):
            if get_position_group(cp) != get_position_group(gp): match = False
        elif fb.get("color") in ("gray","wrong"):
            if get_position_group(cp) == get_position_group(gp): match = False
        if not match: continue

        # Major冠军
        fb = feedbacks.get("major_wins", {})
        gw = guess_player.get("major_wins", 0) or 0
        cw = cand.get("major_wins", 0) or 0
        if fb.get("color") == "green":
            if cw != gw: match = False
        elif fb.get("color") in ("yellow","close"):
            if abs(cw - gw) > 1: match = False
        elif fb.get("color") in ("gray","wrong"):
            if abs(cw - gw) <= 1: match = False
        if match and fb.get("direction"):
            if fb["direction"] == "higher" and cw >= gw: match = False
            if fb["direction"] == "lower" and cw <= gw: match = False
        if not match: continue

        # Major次数
        fb = feedbacks.get("major_appearances", {})
        gma = guess_player.get("major_appearances", 0) or 0
        cma = cand.get("major_appearances", 0) or 0
        if fb.get("color") == "green":
            if cma != gma: match = False
        elif fb.get("color") in ("yellow","close"):
            if abs(cma - gma) > 2: match = False
        elif fb.get("color") in ("gray","wrong"):
            if abs(cma - gma) <= 2: match = False
        if match and fb.get("direction"):
            if fb["direction"] == "higher" and cma >= gma: match = False
            if fb["direction"] == "lower" and cma <= gma: match = False
        if not match: continue

        # 状态
        fb = feedbacks.get("status", {})
        gs = (guess_player.get("status","") or "")
        cs = (cand.get("status","") or "")
        if fb.get("color") == "green":
            if cs.lower() != gs.lower(): match = False
        elif fb.get("color") in ("gray","wrong"):
            if cs.lower() == gs.lower(): match = False

        if match:
            filtered.append(cand)
    return filtered


def recommend_guess(candidates, guessed_names):
    available = [c for c in candidates
                 if c["nickname"].lower() not in guessed_names]
    if not available:
        return None
    if len(available) == 1:
        return available[0]
    return max(available,
               key=lambda c: (c.get("rating", 0) or 0) * 10 + (c.get("major_wins", 0) or 0) * 5)


# ============ 网络拦截 - 核心改进 ============
class NetworkTracker:
    """拦截所有 API 响应，捕获玩家数据和答案"""

    def __init__(self, page: Page):
        self.page = page
        self.api_responses = []  # 所有 API 响应
        self.game_init_data = None  # 游戏初始化数据
        self.answer_data = None  # 答案选手数据
        self.player_list_cache = None  # 玩家列表缓存

        # 监听响应
        page.on("response", self._on_response)

    async def _on_response(self, response):
        url = response.url
        try:
            ct = response.headers.get("content-type", "")
            if "json" not in ct and "/api/" not in url:
                return

            status = response.status
            body = await response.text()

            entry = {
                "url": url,
                "status": status,
                "body_snip": body[:2000],
                "time": time.time(),
            }

            # 尝试解析 JSON
            try:
                data = json.loads(body)
                entry["parsed"] = True

                # 提取可能的答案数据
                if isinstance(data, dict):
                    # 检查是否包含答案
                    if "answer" in data or "solution" in data or "correctPlayer" in data:
                        ans_key = next((k for k in ["answer","solution","correctPlayer","correct_player"]
                                        if k in data), None)
                        if ans_key:
                            ans_val = data[ans_key]
                            if isinstance(ans_val, dict) and "nickname" in ans_val:
                                self.answer_data = ans_val
                                print(f"  [网络] 捕获答案: {ans_val.get('nickname')}")
                            elif isinstance(ans_val, str):
                                self.answer_data = {"nickname": ans_val}
                                print(f"  [网络] 捕获答案昵称: {ans_val}")

                    # 检查是否包含游戏状态 (可能含答案)
                    if "game" in data or "round" in data or "target" in data:
                        self.game_init_data = data
                        if "target" in data and isinstance(data["target"], dict):
                            self.answer_data = data["target"]
                            print(f"  [网络] 从游戏状态捕获答案: {data['target'].get('nickname','?')}")

                    # 玩家列表 API
                    if "/players/list" in url or "/players" in url:
                        if isinstance(data, list):
                            self.player_list_cache = data

                # 检查 list 类型的响应
                if isinstance(data, list) and self.player_list_cache is None:
                    if data and isinstance(data[0], (dict, str)):
                        self.player_list_cache = data

            except json.JSONDecodeError:
                entry["parsed"] = False

            self.api_responses.append(entry)

        except Exception:
            pass

    def get_answer_nickname(self) -> Optional[str]:
        """从拦截的数据中提取答案昵称"""
        if self.answer_data:
            nick = self.answer_data.get("nickname", "")
            if nick: return nick
        return None

    def get_answer_info(self) -> Optional[dict]:
        """从拦截的数据中提取完整答案信息"""
        if self.answer_data and isinstance(self.answer_data, dict):
            nick = self.answer_data.get("nickname") or self.answer_data.get("name")
            if not nick:
                return None
            return {
                "nickname": nick,
                "team": self.answer_data.get("team") or self.answer_data.get("team_name", "Unknown"),
                "country": self.answer_data.get("country") or self.answer_data.get("nationality", "Unknown"),
                "age": self.answer_data.get("age", 0),
                "position": self.answer_data.get("position") or self.answer_data.get("role", "Rifler"),
                "major_wins": self.answer_data.get("major_wins") or self.answer_data.get("majorChampionships", 0),
                "major_appearances": self.answer_data.get("major_appearances") or self.answer_data.get("majorAppearances", 0),
                "status": self.answer_data.get("status") or self.answer_data.get("isActive", "Active"),
                "rating": self.answer_data.get("rating", 0),
            }
        return None

    def dump_for_debug(self):
        """打印所有拦截的 API 用于调试"""
        print("\n  [调试] 拦截的 API 响应:")
        for i, entry in enumerate(self.api_responses):
            body = entry["body_snip"][:300]
            print(f"    [{i}] {entry['status']} {entry['url']}")
            print(f"         {body}")
        print()


# ============ 主类 ============
class AutoPlayer:
    def __init__(self, max_rounds: int = 10):
        self.max_rounds = max_rounds
        self.round_results = []
        self.browser = None
        self.page = None
        self.tracker = None
        self.db_players = []
        self.total_guesses = 0
        self.correct_guesses = 0
        self.known_answer_players = set()  # 已经知道答案的选手（本轮内）

    async def start(self):
        print("\n" + "=" * 65)
        print("  CS2 Guessr 自动游戏 (v2 - 网络拦截版)")
        print(f"  目标轮数: {self.max_rounds}")
        print(f"  数据库已有: {get_player_count()} 位选手")
        print("  浏览器模式: 可视化")
        print("  答案获取: 网络 API 拦截 + DOM 兜底")
        print("=" * 65 + "\n")

        init_db()
        self.db_players = load_all_players()
        export_to_excel()  # 启动时生成初始 Excel

        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(
            headless=False,
            args=["--start-maximized"]
        )
        context = await self.browser.new_context(
            viewport={"width": 1920, "height": 1080},
            no_viewport=False,
        )
        self.page = await context.new_page()

        # 初始化网络拦截器
        self.tracker = NetworkTracker(self.page)

    async def close(self):
        if self.browser:
            await self.browser.close()
        if hasattr(self, 'playwright'):
            await self.playwright.stop()

    async def run(self):
        await self.start()

        print("[1/2] 打开游戏页面...")
        await self.page.goto("https://shnlfriberg.online/single",
                             wait_until="networkidle", timeout=30000)
        await self.page.wait_for_timeout(2000)

        # 关闭公告
        await self._close_notice()

        # 点击"开始游戏"
        print("[2/2] 开始游戏循环...\n")
        await self._click_start()

        for round_idx in range(1, self.max_rounds + 1):
            print(f"{'='*65}")
            print(f"  第 {round_idx}/{self.max_rounds} 轮 | 数据库: {get_player_count()} 位选手")
            print(f"{'='*65}")

            # 重置网络追踪
            self.tracker.api_responses.clear()
            self.tracker.answer_data = None

            result = await self._play_round(round_idx)
            self.round_results.append(result)

            if round_idx < self.max_rounds:
                print(f"\n  准备下一轮...")
                await self._restart_game()

        self._print_summary()
        await self.close()

    async def _play_round(self, round_idx: int) -> dict:
        self.db_players = load_all_players()
        candidates = list(self.db_players)
        guessed_names = set()
        answer_found = None
        all_feedbacks = []

        for guess_num in range(1, 9):
            recommendation = recommend_guess(candidates, guessed_names)
            if not recommendation:
                print(f"  [警告] 无候选选手可猜！")
                # 从外部来源尝试
                for nick in self.known_answer_players:
                    if nick.lower() not in guessed_names:
                        recommendation = {"nickname": nick, "team": "?", "country": "?",
                                         "age": 0, "position": "?", "major_wins": 0,
                                         "major_appearances": 0, "status": "?", "rating": 0}
                        break
                if not recommendation:
                    break

            guess_name = recommendation["nickname"]
            print(f"\n  [{guess_num}/8] 猜测: {guess_name}  |  候选: {len(candidates)} 位")

            success = await self._input_guess(guess_name)
            if not success:
                print(f"    [错误] 输入失败，跳过本轮")
                # 截图调试
                await self._screenshot(f"input_fail_r{round_idx}_g{guess_num}.png")
                break

            feedback = await self._parse_feedback(guess_name)
            if not feedback:
                print(f"    [错误] 无法解析反馈")
                await self._screenshot(f"feedback_fail_r{round_idx}_g{guess_num}.png")
                break

            all_feedbacks.append({"guess": guess_name, "feedback": feedback})
            self._print_feedback(feedback)

            # 检查全部正确
            all_green = all(
                fb.get("color") == "green"
                for k, fb in feedback.items()
                if k != "nickname" and fb.get("color") != "none"
            )
            if all_green:
                print(f"\n  *** 猜对了！答案就是 {guess_name}！***")
                answer_found = guess_name
                self.correct_guesses += 1
                self.known_answer_players.add(guess_name.lower())
                guessed_names.add(guess_name.lower())
                break

            guessed_names.add(guess_name.lower())

            # 构建猜测选手数据
            guess_full = None
            for p in self.db_players:
                if p["nickname"].lower() == guess_name.lower():
                    guess_full = p
                    break

            if guess_full is None:
                guess_full = {
                    "nickname": guess_name,
                    "team": str(feedback.get("team", {}).get("guess_value", "?")),
                    "country": str(feedback.get("country", {}).get("guess_value", "?")),
                    "age": int(feedback.get("age", {}).get("guess_value", 0) or 0),
                    "position": str(feedback.get("position", {}).get("guess_value", "?")),
                    "major_wins": int(feedback.get("major_wins", {}).get("guess_value", 0) or 0),
                    "major_appearances": int(feedback.get("major_appearances", {}).get("guess_value", 0) or 0),
                    "status": str(feedback.get("status", {}).get("guess_value", "?")),
                    "rating": 0,
                }

            before = len(candidates)
            candidates = filter_candidates(candidates, guess_full, feedback)
            removed = before - len(candidates)
            print(f"    排除 {removed} 位，剩余 {len(candidates)} 位候选")

            if len(candidates) == 0:
                print(f"    [警告] 候选清空！重置为排除已猜选手的全部候选")
                candidates = [p for p in self.db_players
                             if p["nickname"].lower() not in guessed_names]
                print(f"    重置为 {len(candidates)} 位")
                if len(candidates) == 0:
                    break

            if len(candidates) <= 5:
                names = ", ".join(c["nickname"] for c in candidates[:5])
                print(f"    候选: [{names}]")

        # 游戏结束 - 使用网络拦截获取答案
        if answer_found is None:
            print(f"\n  未能在 8 次内猜对，尝试获取答案...")
            answer_found = await self._reveal_and_get_answer()

        # 存储答案到数据库
        if answer_found and answer_found != "unknown":
            answer_info = await self._resolve_answer_info(answer_found)
            if answer_info:
                saved = add_player(
                    nickname=answer_info.get("nickname", answer_found),
                    team=answer_info.get("team", "Unknown"),
                    country=answer_info.get("country", "Unknown"),
                    age=answer_info.get("age", 0),
                    position=answer_info.get("position", "Rifler"),
                    major_wins=answer_info.get("major_wins", 0),
                    major_appearances=answer_info.get("major_appearances", 0),
                    status=answer_info.get("status", "Active"),
                    rating=answer_info.get("rating", 0),
                    source="auto",
                )
                if saved:
                    print(f"  [数据库] ✅ 已保存: {answer_info.get('nickname', answer_found)}")
                    self.known_answer_players.add(answer_info.get("nickname", answer_found).lower())
            else:
                print(f"  [数据库] ⚠️ 无法获取 {answer_found} 的详细信息，只存昵称")
                add_player(
                    nickname=answer_found, team="Unknown", country="Unknown",
                    age=0, position="Rifler", major_wins=0, major_appearances=0,
                    status="Active", rating=0, source="auto",
                )

        self.total_guesses += len(all_feedbacks)

        return {
            "round": round_idx,
            "answer": answer_found or "unknown",
            "guesses": len(all_feedbacks),
            "found": all_feedbacks and answer_found == all_feedbacks[-1]["guess"] if all_feedbacks else False,
        }

    # ---- 浏览器操作 ----

    async def _screenshot(self, name: str):
        try:
            path = os.path.join(SCREENSHOT_DIR, name)
            await self.page.screenshot(path=path, full_page=True)
        except:
            pass

    async def _close_notice(self):
        selectors = [
            'button:has-text("我已知晓")',
            'button:has-text("关闭")',
            'button:has-text("知道了")',
            'button:has-text("OK")',
            'button:has-text("确认")',
            'button:has-text("确定")',
        ]
        for sel in selectors:
            try:
                btn = self.page.locator(sel).first
                if await btn.count() > 0 and await btn.is_visible():
                    await btn.click()
                    await self.page.wait_for_timeout(800)
                    print("  已关闭公告")
                    return
            except:
                pass
        try:
            await self.page.keyboard.press("Escape")
            await self.page.wait_for_timeout(500)
        except:
            pass

    async def _click_start(self):
        await self.page.wait_for_timeout(1000)
        start_sels = [
            'button:has-text("开始游戏")',
            'button:has-text("开始")',
            'button:has-text("Start")',
            'button:has-text("Play")',
            'button:has-text("单机模式")',
        ]

        clicked = False
        for sel in start_sels:
            try:
                btn = self.page.locator(sel).first
                await btn.wait_for(state="visible", timeout=5000)
                await btn.click()
                await self.page.wait_for_timeout(3000)
                clicked = True
                print(f"  已点击: {sel}")
                break
            except:
                continue

        if not clicked:
            # 截图调试
            await self._screenshot("start_fail.png")
            # 最后尝试：重新导航到游戏页面
            print(f"  [警告] 未找到开始按钮，尝试重新导航...")
            await self.page.goto("https://shnlfriberg.online/single",
                                 wait_until="networkidle", timeout=30000)
            await self.page.wait_for_timeout(3000)
            await self._close_notice()
            for sel in start_sels:
                try:
                    btn = self.page.locator(sel).first
                    await btn.wait_for(state="visible", timeout=5000)
                    await btn.click()
                    await self.page.wait_for_timeout(3000)
                    clicked = True
                    print(f"  重新导航后已点击: {sel}")
                    break
                except:
                    continue

        if not clicked:
            raise Exception("无法找到开始游戏按钮")

        try:
            await self._wait_for_input_ready()
        except:
            print("  [警告] 等待输入框超时，但继续执行")

    async def _wait_for_input_ready(self):
        try:
            await self.page.wait_for_selector(
                'input[placeholder*="选手"], input.input, input[autocomplete="off"]',
                state="visible", timeout=15000)
            await self.page.wait_for_timeout(1000)
        except:
            pass

    async def _restart_game(self):
        """
        点击"再来一把"开始新一轮。处理可能存在的弹窗遮罩。
        """
        # 先尝试关闭可能残留的弹窗
        await self._dismiss_any_modal()

        restart_sels = [
            'button:has-text("再来一把")',
            'button:has-text("再来一局")',
            'button:has-text("下一轮")',
            'button:has-text("继续游戏")',
            'button:has-text("重新开始")',
            'button:has-text("再来")',
            'button:has-text("Play Again")',
            'button:has-text("Next Round")',
        ]

        clicked = False
        for sel in restart_sels:
            try:
                btn = self.page.locator(sel).first
                await btn.wait_for(state="visible", timeout=5000)
                # force=True 穿透可能的遮罩层
                await btn.click(force=True, timeout=5000)
                await self.page.wait_for_timeout(2500)
                clicked = True
                print(f"  已点击: {sel}")
                break
            except:
                continue

        if not clicked:
            print(f"  [警告] 未找到重新开始按钮，尝试刷新页面...")
            await self._screenshot("restart_fail.png")
            await self.page.goto("https://shnlfriberg.online/single",
                                 wait_until="networkidle", timeout=30000)
            await self.page.wait_for_timeout(3000)
            await self._close_notice()
            await self._click_start()
            return

        # 点击后等待输入框出现
        try:
            await self._wait_for_input_ready()
            print("  已准备就绪")
        except:
            print("  等待输入框超时，继续尝试...")

    async def _dismiss_any_modal(self):
        """关闭页面上任何打开的弹窗/遮罩"""
        # 方法1: 按 Escape
        try:
            await self.page.keyboard.press("Escape")
            await self.page.wait_for_timeout(500)
        except:
            pass

        # 方法2: 点击 backdrop 遮罩
        try:
            backdrop = self.page.locator('.confirm-backdrop, .modal-backdrop, [class*="backdrop"]').first
            if await backdrop.count() > 0 and await backdrop.is_visible():
                await backdrop.click(force=True, timeout=2000)
                await self.page.wait_for_timeout(800)
                print("  已关闭弹窗遮罩")
        except:
            pass

        # 方法3: 点击关闭/取消按钮
        close_sels = [
            'button:has-text("关闭")',
            'button:has-text("取消")',
            'button:has-text("确定")',
            '[aria-label="Close"]',
            '.btn-close',
            '.modal-close',
        ]
        for sel in close_sels:
            try:
                btn = self.page.locator(sel).first
                if await btn.count() > 0 and await btn.is_visible():
                    await btn.click(force=True, timeout=2000)
                    await self.page.wait_for_timeout(500)
                    print(f"  已关闭弹窗: {sel}")
                    return
            except:
                pass

    async def _input_guess(self, nickname: str) -> bool:
        try:
            input_el = self.page.locator(
                'input[role="combobox"], input.input[autocomplete="off"]'
            ).first
            await input_el.wait_for(state="visible", timeout=5000)

            await input_el.click()
            await self.page.wait_for_timeout(300)
            await input_el.fill("")
            await self.page.wait_for_timeout(200)

            await input_el.type(nickname, delay=80)
            await self.page.wait_for_timeout(1500)

            # 等待下拉选项
            try:
                option = self.page.locator('[role="option"]').first
                await option.wait_for(state="visible", timeout=3000)
                await self.page.wait_for_timeout(300)
                await self.page.keyboard.press("Enter")
                await self.page.wait_for_timeout(1000)
            except:
                await self.page.keyboard.press("Enter")
                await self.page.wait_for_timeout(1000)

            # 点击提交按钮
            submit_btn = self.page.locator(
                'form.input-bar button.btn, button:has-text("提交")'
            ).first
            try:
                await submit_btn.wait_for(state="visible", timeout=3000)
                if not await submit_btn.is_disabled():
                    await submit_btn.click()
                else:
                    await self.page.keyboard.press("Enter")
            except:
                await self.page.keyboard.press("Enter")

            await self.page.wait_for_timeout(3000)
            return True
        except Exception as e:
            print(f"    [输入错误] {e}")
            return False

    async def _parse_feedback(self, guess_name: str) -> Optional[dict]:
        try:
            await self.page.wait_for_selector(
                'tr.row-latest, table.game-table tbody tr',
                state="visible", timeout=10000)
            await self.page.wait_for_timeout(500)

            rows = self.page.locator('table.game-table tbody tr')
            row_count = await rows.count()
            if row_count == 0:
                print("    [解析] 未找到反馈行")
                return None

            last_row = rows.nth(row_count - 1)
            cells = last_row.locator('td')
            cell_count = await cells.count()

            if cell_count < 8:
                print(f"    [解析] 行只有 {cell_count} 列")
                return None

            col_attrs = [
                ("team", 1, False),
                ("country", 2, False),
                ("age", 3, True),
                ("position", 4, False),
                ("major_wins", 5, True),
                ("major_appearances", 6, True),
                ("status", 7, False),
            ]

            feedback = {}
            for attr_name, col_idx, is_numeric in col_attrs:
                cell = cells.nth(col_idx)
                class_str = await cell.evaluate("el => el.className")
                classes = class_str.split() if class_str else []

                color = "gray"
                if "correct" in classes or "green" in classes:
                    color = "green"
                elif "close" in classes or "yellow" in classes or "partial" in classes:
                    color = "yellow"
                elif "wrong" in classes:
                    color = "gray"

                cell_text = await cell.evaluate("""
                    el => {
                        let clone = el.cloneNode(true);
                        let dirs = clone.querySelectorAll('.dir, svg, [class*="arrow"]');
                        dirs.forEach(d => d.remove());
                        return clone.textContent.trim();
                    }
                """)

                direction = None
                try:
                    arrow_el = cell.locator('.dir svg, svg.lucide-arrow-up, svg.lucide-arrow-down, [class*="arrow-up"], [class*="arrow-down"]')
                    if await arrow_el.count() > 0:
                        arrow_info = await arrow_el.first.evaluate(
                            "el => ({ class: el.className?.baseVal || el.className || '', outer: el.outerHTML?.substring(0,100) || '' })"
                        )
                        arrow_str = json.dumps(arrow_info).lower()
                        if "arrow-up" in arrow_str or "up" in arrow_str:
                            direction = "lower"
                        elif "arrow-down" in arrow_str or "down" in arrow_str:
                            direction = "higher"
                except:
                    pass

                guess_value = cell_text
                if is_numeric:
                    try:
                        guess_value = int(re.sub(r'[^\d]', '', cell_text.split()[0]))
                    except:
                        guess_value = 0

                feedback[attr_name] = {
                    "color": color,
                    "direction": direction,
                    "guess_value": guess_value,
                }

            # 昵称列
            try:
                name_cell = cells.nth(0)
                name_cls = (await name_cell.evaluate("el => el.className")).split()
                if "correct" in name_cls or "green" in name_cls:
                    feedback["nickname"] = {"color": "green"}
                else:
                    feedback["nickname"] = {"color": "wrong"}
            except:
                feedback["nickname"] = {"color": "none"}

            return feedback

        except Exception as e:
            print(f"    [解析错误] {e}")
            import traceback
            traceback.print_exc()
            return None

    def _print_feedback(self, feedback):
        color_map = {"green": "🟢", "yellow": "🟡", "gray": "⚪", "none": "❓"}
        arrow_map = {"higher": "↑", "lower": "↓", None: " "}
        parts = []
        for attr in ["team","country","age","position","major_wins","major_appearances","status"]:
            fb = feedback.get(attr, {})
            c = color_map.get(fb.get("color","gray"), "?")
            a = arrow_map.get(fb.get("direction"), " ")
            parts.append(f"{c}{a}")
        print(f"    反馈: {' '.join(parts)}")
        vals = []
        for attr in ["team","country","age","position","major_wins","major_appearances","status"]:
            fb = feedback.get(attr, {})
            v = fb.get("guess_value","?")
            vals.append(f"{attr[:3]}={v}")
        print(f"    详情: {' | '.join(vals)}")

    # ---- 答案获取 (改进版) ----

    async def _get_game_id(self) -> Optional[str]:
        """从拦截的 API URL 中提取 gameId"""
        for entry in self.tracker.api_responses:
            m = re.search(r'/game/([a-f0-9-]{36})', entry.get("url", ""))
            if m:
                return m.group(1)
        return None

    async def _reveal_and_get_answer(self) -> Optional[str]:
        """
        获取答案的策略：
        1. 从网络拦截的 API 响应中获取
        2. 点击"查看答案"→弹窗内再点一次→解析 DOM
        3. 尝试直接请求答案 API
        4. 截图保存供调试
        """
        # 策略1: 检查网络拦截
        ans = self.tracker.get_answer_nickname()
        if ans:
            print(f"  [答案] 从网络拦截获取: {ans}")
            return ans

        # 策略2: 点击"查看答案"→等待弹窗→再次点击"查看答案"
        try:
            # 第一步：点击页面上的"查看答案"
            answer_btn_selectors = [
                'button:has-text("查看答案")',
                'button:has-text("显示答案")',
                'button:has-text("答案")',
                'button:has-text("揭晓")',
                'button:has-text("Reveal")',
                'button:has-text("Show Answer")',
            ]

            clicked_first = False
            for sel in answer_btn_selectors:
                try:
                    btn = self.page.locator(sel).first
                    if await btn.count() > 0 and await btn.is_visible():
                        await btn.click()
                        await self.page.wait_for_timeout(2000)
                        clicked_first = True
                        print(f"  已点击: {sel}")
                        break
                except:
                    continue

            if not clicked_first:
                print(f"  [答案] 未找到'查看答案'按钮")
                await self._screenshot("no_answer_btn.png")
                self.tracker.dump_for_debug()
                return "unknown"

            # 第二步：弹窗出现后，在弹窗内再次点击"查看答案"
            print(f"  等待弹窗并再次点击查看答案...")
            await self.page.wait_for_timeout(1500)

            # 弹窗内的按钮选择器（包括 modal/dialog 内的按钮）
            modal_answer_selectors = [
                '.modal button:has-text("查看答案")',
                '[role="dialog"] button:has-text("查看答案")',
                '.confirm-backdrop button:has-text("查看答案")',
                '[class*="modal"] button:has-text("查看答案")',
                '[class*="dialog"] button:has-text("查看答案")',
                'button:has-text("查看答案")',  # 兜底
            ]

            clicked_second = False
            for sel in modal_answer_selectors:
                try:
                    btn = self.page.locator(sel).first
                    if await btn.count() > 0 and await btn.is_visible():
                        await btn.click()
                        await self.page.wait_for_timeout(3000)
                        clicked_second = True
                        print(f"  弹窗内已点击: {sel}")
                        break
                except:
                    continue

            if not clicked_second:
                print(f"  [答案] 弹窗内未找到确认按钮")
                await self._screenshot("no_modal_answer.png")
                self.tracker.dump_for_debug()
                return "unknown"

            # 第三步：答案应该已显示，提取答案
            await self.page.wait_for_timeout(2000)

            # 3a: 检查网络拦截（答案 API 可能刚被调用）
            ans = self.tracker.get_answer_nickname()
            if ans:
                print(f"  [答案] 弹窗确认后从网络获取: {ans}")
                return ans

            # 3b: 从 DOM 提取答案（高亮的选手信息）
            answer_nick = await self.page.evaluate("""
                () => {
                    // 搜索答案相关的高亮元素
                    const selectors = [
                        '[class*="answer"]',
                        '[class*="correct"]',
                        '[class*="solution"]',
                        '[class*="reveal"]',
                        '[class*="highlight"]',
                        '[class*="target"]',
                        '.text-success',
                        '.text-green',
                        '[class*="green"]',
                    ];
                    for (let cls of selectors) {
                        let els = document.querySelectorAll(cls);
                        for (let el of els) {
                            let text = el.textContent.trim();
                            if (text.length > 2 && text.length < 50 && !text.includes('\\n')) {
                                return text;
                            }
                        }
                    }
                    // 搜索整页文本中的答案模式
                    let body = document.body.innerText;
                    let patterns = [
                        /答案[是为：:是]\\s*(\\S{2,30})/,
                        /正确[答案选手]\\S*[是为：:]\\s*(\\S{2,30})/,
                        /Answer[:\\s]+(\\S{2,30})/i,
                        /今[日天]\\S*选手[是为：:]*\\s*(\\S{2,30})/,
                        /correct\\s+player[\\s:]+(\\S{2,30})/i,
                    ];
                    for (let pat of patterns) {
                        let m = body.match(pat);
                        if (m) return m[1];
                    }
                    return null;
                }
            """)
            if answer_nick:
                print(f"  [答案] 从 DOM 提取: {answer_nick}")
                return answer_nick

            # 3c: 尝试直接请求答案 API
            game_id = await self._get_game_id()
            if game_id:
                answer_api_urls = [
                    f"https://shnlfriberg.online/api/game/{game_id}",
                    f"https://shnlfriberg.online/api/game/{game_id}/answer",
                    f"https://shnlfriberg.online/api/game/{game_id}/solution",
                    f"https://shnlfriberg.online/api/game/{game_id}/reveal",
                ]
                for api_url in answer_api_urls:
                    try:
                        resp = await self.page.request.get(api_url)
                        if resp.status == 200:
                            data = await resp.json()
                            print(f"  [API] 尝试 {api_url}: {json.dumps(data, ensure_ascii=False)[:300]}")
                            if isinstance(data, dict):
                                for key in ["answer", "correctPlayer", "solution", "target", "player", "nickname"]:
                                    val = data.get(key)
                                    if isinstance(val, dict) and val.get("nickname"):
                                        return val["nickname"]
                                    if isinstance(val, str) and len(val) > 1:
                                        return val
                                # 直接搜索所有字段中的 nickname
                                for k, v in data.items():
                                    if isinstance(v, dict) and v.get("nickname"):
                                        return v["nickname"]
                    except:
                        continue

            # 兜底：截图保存供调试
            await self._screenshot("reveal_answer.png")
            print(f"  [答案] 已截图 reveal_answer.png 供调试")
            self.tracker.dump_for_debug()

            return "unknown"

        except Exception as e:
            print(f"  [答案-异常] {e}")
            import traceback; traceback.print_exc()
            await self._screenshot("reveal_error.png")
            self.tracker.dump_for_debug()
            return "unknown"

    async def _resolve_answer_info(self, nickname: str) -> Optional[dict]:
        """解析答案选手的完整信息"""
        if not nickname or nickname == "unknown":
            return None

        # 1. 从网络拦截获取
        info = self.tracker.get_answer_info()
        if info and info.get("nickname"):
            return info

        # 2. 查本地数据库
        for p in self.db_players:
            if p["nickname"].lower() == nickname.lower():
                return p

        # 3. 从玩家列表 API 获取
        if self.tracker.player_list_cache:
            for item in self.tracker.player_list_cache:
                if isinstance(item, dict):
                    item_nick = item.get("nickname") or item.get("name") or ""
                    if item_nick.lower() == nickname.lower():
                        return {
                            "nickname": item_nick,
                            "team": item.get("team") or item.get("team_name", "Unknown"),
                            "country": item.get("country") or item.get("nationality", "Unknown"),
                            "age": item.get("age", 0),
                            "position": item.get("position") or item.get("role", "Rifler"),
                            "major_wins": item.get("major_wins") or item.get("majorChampionships", 0),
                            "major_appearances": item.get("major_appearances") or item.get("majorAppearances", 0),
                            "status": item.get("status") or item.get("isActive", "Active"),
                            "rating": item.get("rating", 0),
                        }

        # 4. 尝试调用玩家列表 API
        try:
            resp = await self.page.request.get("https://shnlfriberg.online/api/players/list")
            if resp.status == 200:
                data = await resp.json()
                if isinstance(data, list):
                    self.tracker.player_list_cache = data
                    for item in data:
                        if isinstance(item, dict):
                            item_nick = item.get("nickname") or item.get("name") or ""
                            if item_nick.lower() == nickname.lower():
                                return {
                                    "nickname": item_nick,
                                    "team": item.get("team") or item.get("team_name", "Unknown"),
                                    "country": item.get("country") or item.get("nationality", "Unknown"),
                                    "age": item.get("age", 0),
                                    "position": item.get("position") or item.get("role", "Rifler"),
                                    "major_wins": item.get("major_wins") or item.get("majorChampionships", 0),
                                    "major_appearances": item.get("major_appearances") or item.get("majorAppearances", 0),
                                    "status": item.get("status") or item.get("isActive", "Active"),
                                    "rating": item.get("rating", 0),
                                }
        except Exception as e:
            print(f"    [API] 获取选手列表失败: {e}")

        # 5. 兜底：只返回昵称
        return {
            "nickname": nickname, "team": "Unknown", "country": "Unknown",
            "age": 0, "position": "Rifler", "major_wins": 0,
            "major_appearances": 0, "status": "Active", "rating": 0,
        }

    # ---- 统计 ----

    def _print_summary(self):
        print("\n")
        print("=" * 65)
        print("  游戏结束 - 统计摘要")
        print("=" * 65)
        print(f"  总轮数:         {len(self.round_results)}")
        print(f"  总猜测次数:     {self.total_guesses}")
        print(f"  猜对轮数:       {self.correct_guesses}")
        if self.round_results:
            print(f"  猜对率:         {self.correct_guesses/len(self.round_results)*100:.1f}%")
        avg = self.total_guesses / len(self.round_results) if self.round_results else 0
        print(f"  平均猜测次数:   {avg:.1f}")
        print(f"  数据库选手数:   {get_player_count()}")
        print("=" * 65 + "\n")

        for r in self.round_results:
            status = "✅ 猜对" if r["found"] else "❌ 查答案"
            print(f"  第{r['round']:2d}轮: 答案={r['answer']:20s} | {r['guesses']}次 | {status}")


# ============ 入口 ============
async def main():
    if len(sys.argv) > 1:
        try:
            rounds = int(sys.argv[1])
        except:
            rounds = None
    else:
        rounds = None

    if rounds is None:
        print("\n" + "=" * 65)
        print("  CS2 Guessr 自动游戏 - v2")
        print("=" * 65)
        print("\n  请输入要运行的轮数:")
        try:
            rounds = int(input("  > ").strip())
        except (ValueError, EOFError):
            print("  输入无效，使用默认值 10")
            rounds = 10

    if rounds <= 0:
        print("  轮数必须 > 0")
        return

    player = AutoPlayer(max_rounds=rounds)
    try:
        await player.run()
    except KeyboardInterrupt:
        print("\n\n  用户中断")
        await player.close()
    except Exception as e:
        print(f"\n  [异常] {e}")
        import traceback
        traceback.print_exc()
        try:
            await player.screenshot(f"fatal_error_{int(time.time())}.png")
        except:
            pass
        try:
            await player.close()
        except:
            pass


if __name__ == "__main__":
    asyncio.run(main())
